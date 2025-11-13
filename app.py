import os
import re
import random
import time
import requests
import pandas as pd
import streamlit as st

from io import BytesIO
from typing import Tuple, Optional
from openpyxl import load_workbook
from platform_config import PLATFORM_CONFIG  

# ---------------------------
# APP CONFIG & PAGE
# ---------------------------
st.set_page_config(page_title="Marketplace Reconciliation Tool", layout="wide")

# ---------------------------
# PLATFORM SELECTION (Sidebar)
# ---------------------------
st.sidebar.header("🌐 Select Platform")
platform = st.sidebar.selectbox("Choose Platform", list(PLATFORM_CONFIG.keys()), index=0)
st.info(f"🛍️ Selected Platform: **{platform}** — using its mapping & column rules.")

# ---------------------------
# GOOGLE SHEET READER + RETRY
# ---------------------------
def Read_Sheet_Data(Sheet_Id: str, Sheet_Name: str, HeaderRowIndex: int):
    """
    Reads a Google Sheet via Apps Script web app that returns JSON rows.
    """
    try:
        url = (
            "https://script.google.com/macros/s/"
            "AKfycbxhTA_CQYXW47T7AY0YDDfCVHvdRDoLPqCcb4wAZbJmBm_GOav0eLtmQaqY2T4k_ifDdw"
            f"/exec?sheetId={Sheet_Id}&sheet={Sheet_Name}&headerRowIndex={HeaderRowIndex}"
        )
        response = requests.get(url)
        if response.status_code == 200:
            df = pd.DataFrame(response.json())
            return df, "✅ Success"
        else:
            return False, f"❌ Failed to retrieve sheet data. Status code: {str(response.status_code)}"
    except Exception as e:
        return False, f"❌ Error while reading Google Sheet data Error: {str(e)}"


def read_sheet_with_retry(sheet_id, sheet_name, header_row_index, EndMarkerCol="", retry_attempts=2, wait_time_range=(20, 30), Sheet_Refrence_Name=""):
    """
    Reads data from the Global Sheet with retry logic.
    """
    for attempt in range(1, retry_attempts + 1):
        dataframe, reader_status = Read_Sheet_Data(
            Sheet_Id=sheet_id,
            Sheet_Name=sheet_name,
            HeaderRowIndex=header_row_index
        )
        if dataframe is not False:
            if EndMarkerCol != "":
                try:
                    last_valid_index = dataframe[EndMarkerCol][::-1].replace("", pd.NA).dropna().index[0]
                    dataframe = dataframe.loc[:last_valid_index]
                except Exception:
                    pass
            return dataframe, reader_status

        if attempt < retry_attempts:
            wait_time = random.randint(*wait_time_range)
            time.sleep(wait_time)
    return None, reader_status  # Return None instead of False for consistency


# ---------------------------
# SESSION STATE DEFAULTS
# ---------------------------
if "USER_MAIN_BYTES" not in st.session_state:
    st.session_state.USER_MAIN_BYTES = None
if "USER_SECOND_BYTES" not in st.session_state:
    st.session_state.USER_SECOND_BYTES = None
if "MAIN_SHEET" not in st.session_state:
    st.session_state.MAIN_SHEET = 0
if "TALLY_DETECTED_START" not in st.session_state:
    st.session_state.TALLY_DETECTED_START = None
if "TALLY_DETECTED_END" not in st.session_state:
    st.session_state.TALLY_DETECTED_END = None
if "USER_SELECTED_START_DATE" not in st.session_state:
    st.session_state.USER_SELECTED_START_DATE = None
if "USER_SELECTED_END_DATE" not in st.session_state:
    st.session_state.USER_SELECTED_END_DATE = None
if "DATE_VALIDATION_PASSED" not in st.session_state:
    st.session_state.DATE_VALIDATION_PASSED = None
if "VALIDATION_CHECKED" not in st.session_state:
    st.session_state.VALIDATION_CHECKED = False


# ---------------------------
# UTILITIES
# ---------------------------
@st.cache_data
def to_excel_buffer(df: pd.DataFrame) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False)
    return bio.getvalue()

def _normalize_name(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).strip().lower())

def standardize_columns(df: Optional[pd.DataFrame], mapping: dict) -> Optional[pd.DataFrame]:
    """
    Rename platform-specific columns to standard names, case/space/punct insensitive.
    Supports mapping keys like "MSKU", "Ending Warehouse Balance", and optionally "Location".
    """
    if df is None:
        return None
    rename_dict = {}
    for std_col, plat_col in mapping.items():
        for c in df.columns:
            if _normalize_name(c) == _normalize_name(plat_col):
                rename_dict[c] = std_col
    df = df.rename(columns=rename_dict)
    return df

def ensure_msku(df: Optional[pd.DataFrame], mapping: Optional[dict] = None) -> Optional[pd.DataFrame]:
    """
    Ensure the dataframe has a column named 'MSKU'.
    1) If already present, do nothing.
    2) If a platform mapping is provided, map the platform MSKU-name to 'MSKU'.
    3) Otherwise, rename the FIRST column to 'MSKU'.
    """
    if df is None or df.empty:
        return df
    if "MSKU" in df.columns:
        return df
    # try platform mapping if provided
    if mapping and "MSKU" in mapping:
        target = mapping["MSKU"]
        for c in df.columns:
            if _normalize_name(c) == _normalize_name(target):
                return df.rename(columns={c: "MSKU"})
    # fallback: force first column to MSKU
    first_col = df.columns[0]
    return df.rename(columns={first_col: "MSKU"})

def move_msku_first(df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    if df is None or "MSKU" not in df.columns:
        return df
    cols = ["MSKU"] + [c for c in df.columns if c != "MSKU"]
    return df[cols]


# ---------------------------
# LOAD MAPPING / COMBO (dynamic per platform)
# ---------------------------
@st.cache_data
def load_mapping(selected_platform):
    try:
        config = PLATFORM_CONFIG[selected_platform]
        result = read_sheet_with_retry(
            sheet_id=config["mapping_sheet_id"],
            sheet_name=config["mapping_sheet_name"],
            Sheet_Refrence_Name=f"{selected_platform} Mapping",
            header_row_index=0
        )
        if result[0] is None or result[0] is False:
            st.error(f"Failed to load {selected_platform} Mapping sheet: {result[1]}")
            return None

        mapping_df = result[0]
        mapping_df = mapping_df.rename(columns=lambda c: str(c).strip())

        # Standardize mapping columns to ("Location", "Tally Warehouse Name")
        # per platform mapping_column_map
        col_map = PLATFORM_CONFIG[selected_platform]["mapping_column_map"]
        rename_dict = {}
        for std_col, plat_col in col_map.items():
            for c in mapping_df.columns:
                if _normalize_name(c) == _normalize_name(plat_col):
                    rename_dict[c] = std_col
        mapping_df = mapping_df.rename(columns=rename_dict)
        return mapping_df
    except Exception as e:
        st.error(f"Error loading Mapping sheet: {str(e)}")
        return None

@st.cache_data
def load_combo_master():
    try:
        result = read_sheet_with_retry(
            sheet_id="1YZ7qaoHSBY3RThvwmZPAojw5_9SZF_TRwrUp_WjbVTI",
            sheet_name="Combo",
            Sheet_Refrence_Name="Combo",
            header_row_index=0
        )
        if result[0] is None or result[0] is False:
            st.error(f"Failed to load Combo sheet: {result[1]}")
            return None
        ComboList_Dataframe = result[0]
        ComboList_Dataframe = ComboList_Dataframe.rename(columns=lambda c: str(c).strip())
        return ComboList_Dataframe
    except Exception as e:
        st.error(f"Error loading Combo sheet: {str(e)}")
        return None


# ---------------------------
# LOAD MAIN FILE
# ---------------------------
def load_main_file():
    bytes_data = st.session_state.USER_MAIN_BYTES
    sheet = st.session_state.MAIN_SHEET
    if bytes_data is None:
        return None
    try:
        bio = BytesIO(bytes_data)
        df = pd.read_excel(bio, sheet_name=sheet)
        df = df.rename(columns=lambda c: str(c).strip())
        return df
    except Exception:
        return None


# ---------------------------
# EXCEL DETECTORS / CLEANERS (Second/Tally)
# ---------------------------
def detect_bold_rows(bytes_data, header_index):
    """Return 0-based row indices (relative to data start) where column A is bold."""
    bold_indices = []
    try:
        bio = BytesIO(bytes_data)
        wb = load_workbook(bio, data_only=True)
        ws = wb.active
        data_start = (header_index or 0) + 2  # openpyxl rows are 1-based; preserve original offset
        for r in ws.iter_rows(min_row=data_start, min_col=1, max_col=1):
            cell = r[0]
            if cell.font and cell.font.bold:
                idx = cell.row - data_start
                if idx >= 0:
                    bold_indices.append(idx)
    except Exception:
        pass
    return bold_indices

def scan_sheet_for_warehouse_name(bytes_data, header_index=None):
    """Scan top-left area for warehouse name tokens from Mapping sheet."""
    try:
        bio = BytesIO(bytes_data)
        df_raw = pd.read_excel(bio, header=None, dtype=str)
    except Exception:
        return None

    map_df = load_mapping(platform)
    if map_df is None:
        return None

    names = set()
    if "Tally Warehouse Name" in map_df.columns:
        names.update(map_df["Tally Warehouse Name"].dropna().astype(str).str.strip())
    if "Location" in map_df.columns:
        names.update(map_df["Location"].dropna().astype(str).str.strip())
    names = {n for n in names if len(n) >= 3}
    if not names:
        return None

    max_r = min((header_index or 25), df_raw.shape[0])
    max_c = min(10, df_raw.shape[1])

    for r in range(max_r):
        for c in range(max_c):
            raw = df_raw.iat[r, c]
            val = str(raw).strip() if pd.notna(raw) else ""
            if not val:
                continue
            for wh in names:
                if val.lower() == wh.lower():
                    return wh
                if re.search(rf"(^|\W){re.escape(wh)}($|\W)", val, flags=re.I):
                    return wh
    return None

def clean_second_file(bytes_data):
    """
    Find merged cell containing 'closing' to determine header row,
    read data block, rename quantity columns, drop bold subtotal rows and 'makeup' MSKUs.
    Returns (df, base_name, detected_warehouse)
    """
    try:
        bio = BytesIO(bytes_data)
        df_raw = pd.read_excel(bio, header=None)
        bio.seek(0)
        wb = load_workbook(bio, data_only=True)
        ws = wb.active
    except Exception:
        return None, None, None

    merged_range = None
    try:
        for rng in ws.merged_cells.ranges:
            found = False
            for row in ws[rng.coord]:
                for cell in row:
                    if isinstance(cell.value, str) and re.search(r"\bclosing\b", cell.value.lower()):
                        merged_range = rng
                        found = True
                        break
                if found:
                    break
            if merged_range:
                break
    except Exception:
        pass

    if not merged_range:
        # fallback: try to detect warehouse name and return that
        return None, None, scan_sheet_for_warehouse_name(bytes_data)

    start_col = merged_range.min_col
    end_col = merged_range.max_col
    header_row = merged_range.min_row + 1  # header row index (1-based)

    try:
        bio.seek(0)
        df = pd.read_excel(bio, header=header_row - 1)
    except Exception:
        return None, None, scan_sheet_for_warehouse_name(bytes_data)

    df.columns = [str(c).strip() for c in df.columns]

    # Determine MSKU-like column
    possible_msku_cols = [c for c in df.columns if re.search(r"sku|product|item|article|code|id", str(c), re.I)]
    msku_col = possible_msku_cols[0] if possible_msku_cols else df.columns[0]
    if msku_col != "MSKU":
        df.rename(columns={msku_col: "MSKU"}, inplace=True)

    # Keep only relevant block columns (from merged start to merged end) + MSKU
    relevant_cols = df.columns[start_col - 1:end_col]
    keep_cols = ["MSKU"] + [c for c in relevant_cols if c in df.columns]
    df = df[keep_cols].copy()

    # Normalize common numeric names
    rename_map = {}
    for col in df.columns:
        if re.search(r"qty|quantity", col, re.I):
            rename_map[col] = "Quantity"
        elif re.search(r"rate", col, re.I):
            rename_map[col] = "Rate"
        elif re.search(r"val|amount", col, re.I):
            rename_map[col] = "Value"
    if rename_map:
        df.rename(columns=rename_map, inplace=True)

    # Drop bold subtotal rows + 'makeup' rows
    bold_idx = detect_bold_rows(bytes_data, header_row - 1)
    makeup_idx = df[df["MSKU"].astype(str).str.contains(r"makeup", case=False, na=False)].index.tolist()
    drop_idx = sorted(set(bold_idx + makeup_idx))
    if drop_idx:
        df.drop(index=drop_idx, inplace=True, errors="ignore")
    df.reset_index(drop=True, inplace=True)

    wh = scan_sheet_for_warehouse_name(bytes_data, header_index=header_row - 1)
    base_name = "Second_File"
    return df, base_name, wh


# ---------------------------
# DATE DETECTION & VALIDATION
# ---------------------------
def detect_date_range_from_second(bytes_data) -> Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]:
    """
    Scan top-left area for date-range texts like '1-Apr-25 to 30-Sep-25'.
    Returns (start_date, end_date) as date objects or (None, None).
    """
    try:
        bio = BytesIO(bytes_data)
        df_raw = pd.read_excel(bio, header=None, dtype=str)
    except Exception:
        return None, None

    date_pattern = r"(\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4})\s*to\s*(\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4})"
    for row in df_raw.iloc[:15, :12].fillna("").values.flatten():
        text = str(row).strip()
        match = re.search(date_pattern, text, flags=re.I)
        if match:
            try:
                start_raw, end_raw = match.groups()
                start_date = pd.to_datetime(start_raw, dayfirst=True, errors="coerce").date()
                end_date = pd.to_datetime(end_raw, dayfirst=True, errors="coerce").date()
                if pd.notna(start_date) and pd.notna(end_date):
                    return start_date, end_date
            except Exception:
                continue
    return None, None

def detect_main_file_date_range(df):
    """Find a date column in main file (column name contains 'date') and return min/max date."""
    if df is None or df.empty:
        return None, None
    possible_date_cols = [c for c in df.columns if re.search(r"date", str(c), re.I)]
    for col in possible_date_cols:
        try:
            dates = pd.to_datetime(df[col], errors="coerce")
            valid = dates.dropna()
            if not valid.empty:
                return valid.min().date(), valid.max().date()
        except Exception:
            continue
    return None, None

def check_date_range_compatibility(main_df, second_bytes):
    """
    Ensures ALL rows in Main file have dates within the detected Tally date range.
    If any row is outside the range, display those rows with an error and stop processing.
    Returns True if all dates are within range, False otherwise.
    """
    second_start, second_end = detect_date_range_from_second(second_bytes)

    if not (second_start and second_end):
        st.warning("⚠️ No valid date range found in Tally file — skipping date validation.")
        return True

    if main_df is None or main_df.empty:
        st.warning("⚠️ Main file is empty or unreadable — skipping date validation.")
        return True

    # 1) Candidate date columns by name
    candidate_cols = [c for c in main_df.columns if re.search(r"date", str(c), re.I)]

    # 2) If none by name, also consider columns that parse to datetimes for many rows
    if not candidate_cols:
        for c in main_df.columns:
            try:
                parsed = pd.to_datetime(main_df[c], errors="coerce")
                non_na = parsed.notna().sum()
                if non_na >= max(5, int(0.2 * len(main_df))):
                    candidate_cols.append(c)
            except Exception:
                continue

    if not candidate_cols:
        st.warning("⚠️ No date-like column found in Main file — skipping date validation.")
        return True

    # Prepare boolean frame of out-of-range flags for each candidate column
    out_flag_df = pd.DataFrame(index=main_df.index)
    invalid_dates_info = {}  # store actual detected date per (index, col)

    for col in candidate_cols:
        try:
            dates = pd.to_datetime(main_df[col], errors="coerce")
        except Exception:
            dates = pd.Series([pd.NaT] * len(main_df), index=main_df.index)

        # mark out-of-range only for rows that have a valid date
        out_mask = dates.notna() & ((dates < pd.Timestamp(second_start)) | (dates > pd.Timestamp(second_end)))
        out_flag_df[col] = out_mask

        for idx in out_mask[out_mask].index:
            invalid_dates_info.setdefault(idx, []).append(
                f"{col}: {dates.loc[idx].date() if not pd.isna(dates.loc[idx]) else dates.loc[idx]}"
            )

    any_offending = out_flag_df.any(axis=1)
    if any_offending.any():
        out_rows = main_df.loc[any_offending].copy()
        out_rows["Invalid Date Columns"] = out_flag_df.loc[any_offending].apply(
            lambda r: ", ".join([c for c, flag in r.items() if flag]), axis=1
        )
        out_rows["Invalid Dates"] = out_rows.index.map(lambda i: "; ".join(invalid_dates_info.get(i, [])))

        st.error(
            f"❌ **DATE VALIDATION FAILED!** {len(out_rows)} row(s) have date(s) outside the Tally range "
            f"**({second_start} → {second_end})**. Please correct these dates before proceeding."
        )
        st.dataframe(out_rows, width="stretch")

        try:
            buf = to_excel_buffer(out_rows.reset_index(drop=True))
            st.download_button(
                "⬇ Download Rows with Invalid Dates (Excel)",
                data=buf,
                file_name=f"Invalid_Dates_Outside_Tally_{second_start}_to_{second_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_invalid_dates"
            )
        except Exception:
            pass

        return False

    st.success(f"✅ All rows are within the Tally date range ({second_start} → {second_end}).")
    return True


# ---------------------------
# MAPPING & COMBO
# ---------------------------
def apply_mapping_and_filter(main_df, detected_warehouse):
    map_df = load_mapping(platform)
    if map_df is None:
        return None, None

    # Normalize possible Location column in main file (via platform column_map if provided)
    col_map = PLATFORM_CONFIG[platform]["column_map"]
    # If platform config contains a mapping for "Location", standardize it first
    if "Location" in col_map:
        main_df = standardize_columns(main_df, {"Location": col_map["Location"]})

    if "Location" not in main_df.columns:
        return None, None

    main_df = main_df.rename(columns=lambda c: str(c).strip())
    map_df = map_df.rename(columns=lambda c: str(c).strip())

    merged = pd.merge(main_df, map_df, on="Location", how="left")

    if detected_warehouse is None:
        return None, None

    def get_tally(map_df, loc):
        if "Location" not in map_df.columns or "Tally Warehouse Name" not in map_df.columns:
            return loc
        match = map_df[
            map_df["Location"].astype(str).str.lower().str.strip() == str(loc).lower().strip()
        ]
        if not match.empty and pd.notna(match.iloc[0]["Tally Warehouse Name"]):
            return match.iloc[0]["Tally Warehouse Name"]
        return loc

    tally = get_tally(map_df, detected_warehouse)
    filtered = merged[merged["Tally Warehouse Name"].astype(str) == str(tally)]
    if filtered.empty:
        return None, None
    return filtered.reset_index(drop=True), tally

def apply_combo_double_replace(df, combo_df):
    if df is None or combo_df is None or combo_df.empty:
        return df

    if "MSKU" not in df.columns:
        return df

    combo = combo_df.copy()
    combo["Parent SKU"] = combo["Parent SKU"].fillna("").str.strip().str.upper()
    combo["Component SKU"] = combo["Component SKU"].fillna("").str.strip().str.upper()
    combo["Product Config"] = combo["Product Config"].fillna("").str.strip().str.lower()

    combo = combo[
        (combo["Parent SKU"] != "")
        & (combo["Component SKU"] != "")
        & (combo["Product Config"].isin(["replacement", "double", "combo"]))
    ]

    grouped = {}
    for _, r in combo.iterrows():
        grouped.setdefault(r["Parent SKU"], []).append((r["Product Config"], r["Component SKU"]))

    new_rows = []
    for _, row in df.iterrows():
        msku = str(row.get("MSKU", "")).strip().upper()
        if not msku or msku not in grouped:
            new_rows.append(row.copy())
            continue

        entries = grouped[msku]
        comps_replace = [c for cfg, c in entries if cfg == "replacement"]
        comps_double  = [c for cfg, c in entries if cfg == "double"]
        comps_combo   = [c for cfg, c in entries if cfg == "combo"]

        if comps_replace:
            for c in comps_replace + comps_combo + comps_double:
                nr = row.copy()
                nr["MSKU"] = c
                new_rows.append(nr)
            continue

        if comps_double:
            for c in comps_double:
                for _ in range(2):
                    nr = row.copy()
                    nr["MSKU"] = c
                    new_rows.append(nr)
            for c in comps_combo:
                nr = row.copy()
                nr["MSKU"] = c
                new_rows.append(nr)
            continue

        if comps_combo:
            for c in comps_combo:
                nr = row.copy()
                nr["MSKU"] = c
                new_rows.append(nr)
            continue

    return pd.DataFrame(new_rows).reset_index(drop=True)


# ---------------------------
# COMBINE & DIFFERENCE
# ---------------------------
def combine(df_main, df_second):
    if df_main is None or df_second is None:
        return pd.DataFrame()
    df_main = df_main.copy()
    df_second = df_second.copy()
    df_main["MSKU"] = df_main["MSKU"].astype(str)
    df_second["MSKU"] = df_second["MSKU"].astype(str)
    combined = pd.merge(df_main, df_second, on="MSKU", how="inner")
    return combined

def create_msku_difference_report(filtered_df, tally_df, tally_label, platform_name):
    if filtered_df is None or tally_df is None:
        return None, None, None
    left = filtered_df.copy()
    right = tally_df.copy()
    left["MSKU"] = left["MSKU"].astype(str).str.strip()
    right["MSKU"] = right["MSKU"].astype(str).str.strip()

    # ✅ Dynamically name quantity column based on selected platform
    end_col = f"{platform_name} Quantity"
    qty_col = "Tally Quantity"

    if "Ending Warehouse Balance" not in left.columns:
        left[end_col] = pd.NA
    else:
        left[end_col] = left["Ending Warehouse Balance"]

    if "Quantity" not in right.columns:
        right[qty_col] = pd.NA
    else:
        right[qty_col] = right["Quantity"]

    left_g = left.groupby("MSKU", as_index=False).agg({end_col: "sum"})
    right_g = right.groupby("MSKU", as_index=False).agg({qty_col: "sum"})

    merged = pd.merge(left_g, right_g, on="MSKU", how="outer").fillna(0)

    merged[qty_col] = pd.to_numeric(merged[qty_col], errors="coerce").fillna(0)
    merged[end_col] = pd.to_numeric(merged[end_col], errors="coerce").fillna(0)
    merged["Difference"] = merged[qty_col] - merged[end_col]

    total = pd.DataFrame({
        "MSKU": ["TOTAL"],
        qty_col: [merged[qty_col].sum()],
        end_col: [merged[end_col].sum()],
        "Difference": [merged["Difference"].sum()]
    })
    merged = pd.concat([merged, total], ignore_index=True)

    out_name = f"{platform_name}_Difference_MSKU_{tally_label}.xlsx"
    return merged, out_name, to_excel_buffer(merged)


# ---------------------------
# MAIN PIPELINE (end-to-end)
# ---------------------------
def run_pipeline(require_containment=False, block_if_out_of_range=False):
    progress_bar = st.progress(0)
    status_text = st.empty()
    steps = [
        "Loading Main File",
        "Filtering Main File (user selected range)",
        "Cleaning Second File",
        "Applying Mapping",
        "Loading Combo Master",
        "Applying Combo Logic (Main)",
        "Applying Combo Logic (Second)",
        "Combining Data",
        "Creating Difference Report",
    ]

    # Step 1: load main
    status_text.text(steps[0])
    progress_bar.progress(0.02)
    main_df = load_main_file()
    if main_df is None:
        st.error("❌ Could not read the Main Excel file (selected sheet). Check file format/headers.")
        return

    # Normalize / ensure MSKU / Location (platform-aware)
    col_map = PLATFORM_CONFIG[platform]["column_map"]
    main_df = standardize_columns(main_df, col_map)
    main_df = ensure_msku(main_df, col_map)
    main_df = move_msku_first(main_df)
    # If platform defines Location alias in column_map, it was handled above.

    # Step 2: filter main by user-selected date range (if set)
    status_text.text(steps[1])
    progress_bar.progress(0.15)
    if st.session_state.USER_SELECTED_START_DATE and st.session_state.USER_SELECTED_END_DATE:
        possible_date_cols = [c for c in main_df.columns if re.search(r"date", str(c), re.I)]
        if possible_date_cols:
            filtered = None
            for col in possible_date_cols:
                try:
                    main_df[col] = pd.to_datetime(main_df[col], errors="coerce")
                    mask = (main_df[col] >= pd.Timestamp(st.session_state.USER_SELECTED_START_DATE)) & \
                           (main_df[col] <= pd.Timestamp(st.session_state.USER_SELECTED_END_DATE))
                    filtered = main_df.loc[mask].copy()
                    if filtered is not None:
                        main_df = filtered.reset_index(drop=True)
                        st.info(f"Filtered Main file by column '{col}' to range {st.session_state.USER_SELECTED_START_DATE} → {st.session_state.USER_SELECTED_END_DATE}. Rows remaining: {len(main_df)}")
                        break
                except Exception:
                    continue

    # Step 3: clean second file
    status_text.text(steps[2])
    progress_bar.progress(0.35)
    df2, second_name, wh = clean_second_file(st.session_state.USER_SECOND_BYTES)
    if df2 is None:
        st.error("❌ Could not parse the Second Excel file or detect the 'closing' block.")
        return

    # Standardize second file columns (platform-aware) + ensure MSKU
    df2 = standardize_columns(df2, col_map)
    df2 = ensure_msku(df2, col_map)
    df2 = move_msku_first(df2)

    # Step 4: mapping
    status_text.text(steps[3])
    progress_bar.progress(0.50)
    filtered_main, tally = apply_mapping_and_filter(main_df, wh)
    if filtered_main is None:
        st.error("❌ Mapping/warehouse filtering failed. Ensure platform Mapping is present and 'Location' exists in main file (after normalization).")
        return

    # Step 5: load combo master
    status_text.text(steps[4])
    progress_bar.progress(0.62)
    combo_master = load_combo_master()
    if combo_master is None:
        st.warning("⚠ Combo master not found or unreadable. Proceeding without combo expansion.")

    # Step 6: apply combo to main
    status_text.text(steps[5])
    progress_bar.progress(0.72)
    combo_main = apply_combo_double_replace(filtered_main, combo_master)

    # Step 7: apply combo to second
    status_text.text(steps[6])
    progress_bar.progress(0.82)
    combo_second = apply_combo_double_replace(df2, combo_master)

    # Ensure MSKU still present after combo (safety)
    if "MSKU" not in combo_main.columns:
        combo_main = ensure_msku(combo_main, col_map)
    if "MSKU" not in combo_second.columns:
        combo_second = ensure_msku(combo_second, col_map)

    # Step 8: combine and diff
    status_text.text(steps[7])
    progress_bar.progress(0.90)
    combined = combine(combo_main, combo_second)
    buf_combined = to_excel_buffer(combined)

    progress_bar.progress(0.95)
    status_text.text(steps[8])
    diff_df, diff_name, diff_buf = create_msku_difference_report(
        combo_main, combo_second, re.sub(r"[^\w\-_]", "_", str(tally)), platform
    )

    if diff_df is None:
        st.error("❌ Failed to create the MSKU difference report.")
        return

    progress_bar.progress(1.0)
    status_text.text("✅ Processing Complete")

    # Display detected info & downloads
    st.subheader("📋 Detected Information")
    s_start, s_end = detect_date_range_from_second(st.session_state.USER_SECOND_BYTES)
    main_start, main_end = detect_main_file_date_range(load_main_file())
    if s_start and s_end:
        st.write(f"Detected Tally Date Range: **{s_start} to {s_end}**")
    else:
        st.write("Detected Tally Date Range: **Not found**")
    if main_start and main_end:
        st.write(f"Main File Date Range: **{main_start} to {main_end}**")
    else:
        st.write("Main File Date Range: **Not found**")

    st.write(f"Detected Warehouse (from second file): **{wh}**")
    st.write(f"Tally Warehouse Name (mapped): **{tally}**")
    st.write(f"Second File Base Name: **{second_name}**")

    st.subheader("📥 Downloads")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            "⬇ Filtered Main (after combo)",
            data=to_excel_buffer(combo_main),
            file_name=f"Filtered_{re.sub(r'[^\w\-_]', '_', str(tally))}_after_combo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.download_button(
            "⬇ Combined (inner on MSKU)",
            data=buf_combined,
            file_name=f"Combined_{re.sub(r'[^\w\-_]', '_', str(tally))}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    with c2:
        st.download_button(
            f"⬇ TALLY {second_name} (after combo)",
            data=to_excel_buffer(combo_second),
            file_name=f"TALLY_{second_name}_after_combo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.download_button(
            "⬇ MSKU Difference Report",
            data=diff_buf,
            file_name=diff_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ---------------------------
# UI
# ---------------------------
st.title("📊 Market Place RECONCILIATION PROJECT")
st.markdown("""
This app:
- Detects bold subtotal rows and removes them.
- Detects warehouse name from second file and maps using platform Mapping.
- Applies replacement/double/combo rules from Combo master.
- Detects Tally date range and lets you select a Main-file date window to filter.
- **Automatically validates dates** when both files are uploaded.
- Works across platforms with **column-name normalization** (MSKU, Ending Warehouse Balance, Location).
- Guarantees **MSKU** exists by renaming the first column if needed.
""")

left_col, right_col = st.columns(2)

# -------------------------------------
# MAIN FILE UPLOAD
# -------------------------------------
with left_col:
    st.subheader("📦 Upload MAIN DATA")
    main_file = st.file_uploader("Select MAIN DATA (.xlsx)", type=["xlsx"], key="main_up")
    if main_file is not None:
        try:
            bio = BytesIO(main_file.getvalue())
            wb = load_workbook(bio, data_only=True)
            sheet_names = wb.sheetnames

            if len(sheet_names) > 1:
                selected = st.selectbox("Select Sheet from MAIN DATA", sheet_names, index=0)
                st.session_state.MAIN_SHEET = sheet_names.index(selected)
            else:
                st.session_state.MAIN_SHEET = 0

            st.session_state.USER_MAIN_BYTES = main_file.getvalue()
            st.session_state.VALIDATION_CHECKED = False

        except Exception:
            st.session_state.MAIN_SHEET = 0
            st.session_state.USER_MAIN_BYTES = main_file.getvalue()
            st.session_state.VALIDATION_CHECKED = False

# -------------------------------------
# SECOND FILE UPLOAD
# -------------------------------------
with right_col:
    st.subheader("🏢 Upload Tally Warehouse DATA")
    second_file = st.file_uploader("Select Tally Warehouse DATA (.xlsx)", type=["xlsx"], key="second_up")
    if second_file is not None:
        st.session_state.USER_SECOND_BYTES = second_file.getvalue()
        st.session_state.VALIDATION_CHECKED = False

        t_start, t_end = detect_date_range_from_second(second_file.getvalue())
        if t_start and t_end:
            st.info(f"📅 Detected Tally Date Range: **{t_start} → {t_end}**")
            st.session_state.TALLY_DETECTED_START = t_start
            st.session_state.TALLY_DETECTED_END = t_end

# -------------------------------------
# DATE VALIDATION (Auto)
# -------------------------------------
st.markdown("---")

if st.session_state.USER_MAIN_BYTES and st.session_state.USER_SECOND_BYTES:

    # First-time validation only
    if not st.session_state.VALIDATION_CHECKED:
        st.subheader("🔍 Automatic Date Validation")

        with st.spinner("Validating dates automatically..."):
            main_df_for_check = load_main_file()

            col_map = PLATFORM_CONFIG[platform]["column_map"]
            main_df_for_check = standardize_columns(main_df_for_check, col_map)
            main_df_for_check = ensure_msku(main_df_for_check, col_map)

            if main_df_for_check is not None:
                validation_passed = check_date_range_compatibility(
                    main_df_for_check, st.session_state.USER_SECOND_BYTES
                )
                st.session_state.DATE_VALIDATION_PASSED = validation_passed
                st.session_state.VALIDATION_CHECKED = True

            else:
                st.error("❌ Could not load Main file for validation.")
                st.session_state.DATE_VALIDATION_PASSED = False
                st.session_state.VALIDATION_CHECKED = True

    else:
        # If already validated
        st.subheader("🔍 Date Validation Status")

        if st.session_state.DATE_VALIDATION_PASSED:
            st.success("✅ Date validation passed! You can proceed with processing.")
        else:
            st.error("❌ Date validation failed. Please fix the dates in your Main file and re-upload.")

    st.markdown("---")

    # -------------------------------------
    # RUN PROCESSING BUTTON (Now shown ONLY here)
    # -------------------------------------
    run_button_disabled = False
    run_button_message = ""

    if not st.session_state.VALIDATION_CHECKED:
        run_button_disabled = True
        run_button_message = "⚠️ Please wait for automatic date validation to complete."

    elif not st.session_state.DATE_VALIDATION_PASSED:
        run_button_disabled = True
        run_button_message = "❌ Date validation failed. Please correct the dates in your Main file and re-upload."

    if run_button_message:
        st.warning(run_button_message)

    run = st.button("🚀 Run Processing", type="primary", disabled=run_button_disabled)

    if run:
        run_pipeline()

st.markdown("---")
